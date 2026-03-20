#!/usr/bin/env python3
"""Wait for USB replug, then immediately run EIS sweep in the same process."""
import serial
import serial.tools.list_ports
import time
import sys
import os

PORT = "COM4"
BAUD = 115200
START_HZ = 10.0
END_HZ = 100000.0
AMPLITUDE_MV = 200.0
TIMEOUT = 600
OUT_FILE = "eis_sweep_data.xlsx"

def P(*args):
    print(*args, flush=True)

def find_esp():
    for p in serial.tools.list_ports.comports():
        if "303A" in p.hwid:
            return p.device
    return None

# Step 1: wait for unplug
P("Waiting for USB cable to be UNPLUGGED...")
while find_esp():
    time.sleep(0.5)
P("Board disconnected! Now PLUG IT BACK IN...")

# Step 2: wait for replug
while not find_esp():
    time.sleep(0.5)
port = find_esp()
P(f"Board found on {port}! Waiting 5 sec for firmware to boot...")
time.sleep(5)

# Step 3: open serial and run sweep
P(f"Opening {port}...")
ser = serial.Serial()
ser.port = port
ser.baudrate = BAUD
ser.timeout = 1
ser.write_timeout = 2
ser.dsrdtr = False
ser.rtscts = False

try:
    ser.open()
except Exception as e:
    P(f"Failed to open port: {e}")
    sys.exit(1)

time.sleep(0.5)

# Read boot messages
P("--- Boot messages ---")
try:
    data = ser.read(4096)
    if data:
        text = data.decode("utf-8", errors="replace").strip()
        for line in text.split("\n"):
            P(f"  {line.strip()}")
        if "ESP-ROM" in text and "rst:0x7" in text:
            P("FATAL: Firmware crashed (WDT). Cannot proceed.")
            ser.close()
            sys.exit(1)
    else:
        P("  (none)")
except Exception as e:
    P(f"  read error: {e}")

# Send EIS command
P()
P(f"Sending EIS sweep: {START_HZ} Hz - {END_HZ} Hz, {AMPLITUDE_MV} mV")
cmd = f"MEASURE:0,{START_HZ},{END_HZ},10,0.0,0.0,1000.0,1,1,127000.0,150.0,0,0,{AMPLITUDE_MV}\n"
try:
    ser.write(cmd.encode("utf-8"))
except Exception as e:
    P(f"Write failed: {e}")
    ser.close()
    sys.exit(1)

P("Command sent. Capturing data...")
P()

rows = []
capturing = False
done = False
start = time.time()

while not done and (time.time() - start < TIMEOUT):
    try:
        line_bytes = bytearray()
        t0 = time.time()
        while time.time() - t0 < 2.0:
            b = ser.read(1)
            if not b:
                if line_bytes:
                    break
                continue
            if b == b'\n':
                break
            line_bytes.extend(b)
    except (serial.SerialException, OSError):
        time.sleep(0.1)
        continue

    raw = bytes(line_bytes).decode("utf-8", errors="ignore").strip()
    if not raw:
        continue

    P(f"  {raw}")

    if "Frequency" in raw and "DFT" in raw:
        capturing = True
        continue

    if capturing:
        if raw.startswith("---") or raw.startswith("Sweep complete") or raw.startswith("Saved") or raw.startswith("=== Measurement"):
            if rows:
                done = True
                break
            continue

        parts = raw.split(",")
        if len(parts) >= 6:
            try:
                int(parts[0])   # Index
                float(parts[1]) # Frequency
                rows.append(raw)
            except ValueError:
                if rows:
                    done = True
                    break

elapsed = time.time() - start
ser.close()
P()
P(f"Capture done. {len(rows)} data points in {elapsed:.1f}s")

if not rows:
    P("No CSV data captured.")
    sys.exit(1)

# Save to Excel
try:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EIS Data"
    headers = ["Index", "Freq(Hz)", "DFT_Cal", "DFT_Mag", "Rz(Ohm)", "Rreal(Ohm)", "Rimag(Ohm)", "Phase(rad)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, 2):
        parts = row.split(",")
        for col, val in enumerate(parts[:8], 1):
            try:
                ws.cell(row=row_idx, column=col, value=float(val))
            except ValueError:
                ws.cell(row=row_idx, column=col, value=val)
    wb.save(OUT_FILE)
    P(f"Saved to: {os.path.abspath(OUT_FILE)}")
except ImportError:
    csv_file = OUT_FILE.replace(".xlsx", ".csv")
    with open(csv_file, "w") as f:
        f.write("Index,Freq(Hz),DFT_Cal,DFT_Mag,Rz(Ohm),Rreal(Ohm),Rimag(Ohm),Phase(rad)\n")
        for row in rows:
            f.write(row + "\n")
    P(f"Saved to: {os.path.abspath(csv_file)}")

P("Done!")
