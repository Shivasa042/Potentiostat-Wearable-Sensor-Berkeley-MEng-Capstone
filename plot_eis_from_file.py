#!/usr/bin/env python3
"""
Plot Nyquist and Bode diagrams from EIS data saved as CSV or Excel.

Supports:
  - Serial monitor / printDataCSV format:
      Frequency(Hz),Real(Ohm),Imaginary(Ohm),Magnitude(Ohm),Phase(Degrees)
  - SD card saveDataEIS format (first cycle columns):
      Freq, Magnitude, Phase (rad), Phase (deg), Real, Imag
  - Minimal export: Frequency,Real,Imag  (or Freq, Real, Imag)

Usage:
  python plot_eis_from_file.py --input sweep.csv
  python plot_eis_from_file.py --input results.xlsx --output nyquist_bode.png

Dependencies: numpy, matplotlib; for .xlsx also openpyxl (see requirements.txt)
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import re
import sys
from typing import List, Optional, Tuple


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _find_col(headers: List[str], patterns: Tuple[str, ...]) -> Optional[int]:
    """Return index of first header matching any pattern (substring, normalized)."""
    nh = [_norm(h) for h in headers]
    for i, h in enumerate(nh):
        for p in patterns:
            if p in h:
                return i
    return None


def _parse_float(x) -> float:
    if x is None or x == "":
        return float("nan")
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        return float(x)
    s = str(x).strip().lower()
    if s in ("inf", "+inf", "infinity"):
        return float("inf")
    if s in ("-inf", "-infinity"):
        return float("-inf")
    if s in ("nan", ""):
        return float("nan")
    try:
        return float(s)
    except ValueError:
        return float("nan")


def load_rows_csv(path: str) -> Tuple[List[str], List[List]]:
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        raise ValueError("CSV file is empty.")
    return rows[0], rows[1:]


def load_rows_xlsx(path: str) -> Tuple[List[str], List[List]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(
            "Reading .xlsx requires openpyxl. Install with: pip install openpyxl"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        raise ValueError("Excel sheet is empty.")
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data = rows[1:]
    return headers, data


def extract_impedance(headers: List[str], data_rows: List[List]) -> Tuple:
    """Return (frequency, real_Z, imag_Z) as lists of finite floats only."""
    # Map flexible column names
    i_freq = _find_col(
        headers,
        ("frequency(hz)", "frequency", "freq"),
    )
    i_real = _find_col(
        headers,
        ("real(ohm)", "real", "z'"),
    )
    i_imag = _find_col(
        headers,
        ("imaginary(ohm)", "imaginary", "imag", "z''", 'z"'),
    )

    if i_freq is None or i_real is None or i_imag is None:
        raise ValueError(
            "Could not find frequency / real / imaginary columns.\n"
            f"Headers found: {headers}\n"
            "Expected names like Frequency(Hz), Real(Ohm), Imaginary(Ohm) "
            "or Freq, Real, Imag."
        )

    freq, real, imag = [], [], []
    for row in data_rows:
        if i_freq >= len(row) or i_real >= len(row) or i_imag >= len(row):
            continue
        f = _parse_float(row[i_freq])
        r = _parse_float(row[i_real])
        im = _parse_float(row[i_imag])
        if math.isfinite(f) and math.isfinite(r) and math.isfinite(im):
            freq.append(f)
            real.append(r)
            imag.append(im)

    if not freq:
        raise ValueError(
            "No finite impedance rows after filtering. "
            "Open-circuit sweeps often produce 'inf' — connect a cell or resistor."
        )
    return freq, real, imag


def plot_eis(freq: List[float], real: List[float], imag: List[float], title: str = ""):
    import numpy as np
    import matplotlib.pyplot as plt

    f = np.array(freq, dtype=np.float64)
    zr = np.array(real, dtype=np.float64)
    zi = np.array(imag, dtype=np.float64)
    mag = np.sqrt(zr**2 + zi**2)
    phase_deg = np.degrees(np.arctan2(zi, zr))

    fig, axs = plt.subplots(1, 2, figsize=(14, 6))

    # Nyquist: Z' vs -Z'' (electrochemical convention)
    axs[0].plot(zr, -zi, "o-", markersize=4)
    axs[0].set_xlabel(r"Real Z ($\Omega$)")
    axs[0].set_ylabel(r"-Imag Z ($\Omega$)")
    axs[0].set_title("Nyquist plot")
    axs[0].grid(True, which="both", alpha=0.3)
    axs[0].set_aspect("equal", adjustable="box")

    # Bode
    ax_mag = axs[1]
    ax_mag.semilogx(f, mag, "g-o", markersize=3, label="|Z|")
    ax_mag.set_xlabel("Frequency (Hz)")
    ax_mag.set_ylabel(r"Magnitude ($\Omega$)", color="g")
    ax_mag.tick_params(axis="y", labelcolor="g")
    ax_mag.grid(True, which="both", alpha=0.3)
    ax_mag.set_title("Bode plot")

    ax_phase = ax_mag.twinx()
    ax_phase.semilogx(f, phase_deg, "r--s", markersize=2, label="Phase")
    ax_phase.set_ylabel("Phase (°)", color="r")
    ax_phase.tick_params(axis="y", labelcolor="r")

    if title:
        fig.suptitle(title)
        plt.tight_layout(rect=[0, 0, 1, 0.96])
    else:
        plt.tight_layout()

    return fig


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Plot Nyquist and Bode from EIS CSV or Excel export."
    )
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        help="Path to .csv or .xlsx (from serial copy-paste, SD card, or eis_csv_to_xlsx)",
    )
    parser.add_argument(
        "--output",
        "-o",
        default="",
        help="Save figure to this file (e.g. nyquist_bode.png). If omitted, only show window.",
    )
    parser.add_argument(
        "--title",
        default="",
        help="Optional plot title",
    )
    parser.add_argument(
        "--no-show",
        action="store_true",
        help="Do not open interactive window (use with --output)",
    )
    args = parser.parse_args()

    path = args.input
    if not os.path.isfile(path):
        print(f"Error: file not found: {path}", file=sys.stderr)
        return 1

    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        headers, data = load_rows_csv(path)
    elif ext in (".xlsx", ".xlsm"):
        headers, data = load_rows_xlsx(path)
    else:
        print(
            "Error: use a .csv or .xlsx file. For .xls, export as .xlsx in Excel.",
            file=sys.stderr,
        )
        return 1

    # Drop completely empty trailing columns in row data
    headers = [h.strip() if isinstance(h, str) else str(h) for h in headers]
    try:
        freq, real, imag = extract_impedance(headers, data)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    print(f"Loaded {len(freq)} finite frequency points from {path}")

    fig = plot_eis(freq, real, imag, title=args.title or os.path.basename(path))

    if args.output:
        fig.savefig(args.output, dpi=150, bbox_inches="tight")
        print(f"Saved figure to {args.output}")

    if args.no_show:
        if not args.output:
            print(
                "Warning: --no-show without --output; figure was not saved or displayed.",
                file=sys.stderr,
            )
    else:
        import matplotlib.pyplot as plt

        plt.show()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
