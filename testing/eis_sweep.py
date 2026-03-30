#!/usr/bin/env python3
"""
All-in-one EIS sweep script for ESP32-S3 with USB Serial/JTAG on Windows.

Workflow:
1. Waits for user to unplug and replug USB (exits bootloader loop)
2. Waits for firmware to boot
3. Opens COM port and sends EIS sweep command
4. Captures data and saves to Excel (.xlsx)
5. Closes port

Uses ser.read() with manual line parsing to avoid Windows ClearCommError issues.
"""
import serial
import serial.tools.list_ports
import time
import sys
import os

def pprint(*args, **kwargs):
    """Print with flush for non-terminal output."""
    kwargs.setdefault("flush", True)
    print(*args, **kwargs)

PORT = "COM4"
BAUD = 115200
START_HZ = 10.0
END_HZ = 100000.0
AMPLITUDE_MV = 200.0
TIMEOUT = 180  # seconds for sweep data capture

OUT_FILE = "eis_sweep_data.xlsx"


def find_esp_port():
    """Find the ESP32-S3 COM port."""
    for p in serial.tools.list_ports.comports():
        if "303A" in p.hwid:
            return p.device
    return None


def robust_read(ser, max_bytes=4096, timeout_s=2.0):
    """Read available data using only ReadFile, no ClearCommError."""
    start = time.time()
    chunks = []
    while time.time() - start < timeout_s:
        try:
            chunk = ser.read(min(max_bytes, 256))
            if chunk:
                chunks.append(chunk)
                start = time.time()  # reset timeout on new data
            else:
                break
        except (serial.SerialException, OSError):
            break
    return b"".join(chunks)


def robust_readline(ser, timeout_s=2.0):
    """Read until newline or timeout, byte by byte to avoid ClearCommError issues."""
    line = bytearray()
    start = time.time()
    while time.time() - start < timeout_s:
        try:
            b = ser.read(1)
            if not b:
                if line:
                    break
                continue
            if b == b'\n':
                break
            line.extend(b)
        except (serial.SerialException, OSError):
            break
    return bytes(line)


def main():
    pprint("=" * 60)
    pprint("ESP32-S3 EIS Sweep — USB Serial/JTAG Edition")
    pprint("=" * 60)
    pprint()

    # Check if port exists
    port = find_esp_port()
    if not port:
        pprint(f"No ESP32-S3 found. Waiting for USB connection...")
        while not port:
            time.sleep(0.5)
            port = find_esp_port()

    pprint(f"Found ESP32-S3 on {port}")
    pprint()

    # Open serial port
    pprint(f"Opening {port}...")
    ser = serial.Serial()
    ser.port = port
    ser.baudrate = BAUD
    ser.timeout = 1
    ser.write_timeout = 2
    ser.dsrdtr = False
    ser.rtscts = False

    try:
        ser.open()
    except serial.SerialException as e:
        pprint(f"Failed to open {port}: {e}")
        pprint("The port might be held by another process or the board needs a USB replug.")
        return 1

    time.sleep(0.5)

    # Read any pending boot messages
    pprint("Reading boot messages...")
    boot_data = robust_read(ser, timeout_s=3.0)
    if boot_data:
        text = boot_data.decode("utf-8", errors="replace").strip()
        pprint(f"  {text[:300]}")
        if "ESP-ROM" in text and "rst:0x7" in text:
            pprint()
            pprint("ERROR: Board is in bootloader mode (firmware crashed).")
            pprint("Please unplug and replug the USB cable, then run this script again.")
            ser.close()
            return 1
    else:
        pprint("  (no pending data)")

    pprint()
    pprint(f"Sending EIS sweep: {START_HZ} Hz - {END_HZ} Hz, {AMPLITUDE_MV} mV")

    command = (
        f"MEASURE:0,{START_HZ},{END_HZ},10,0.0,0.0,1000.0,1,1,"
        f"127000.0,150.0,0,0,{AMPLITUDE_MV}\n"
    )

    try:
        ser.write(command.encode("utf-8"))
    except (serial.SerialException, OSError) as e:
        pprint(f"Write failed: {e}")
        ser.close()
        return 1

    pprint("Command sent. Waiting for data...")
    pprint()

    # Capture output
    rows = []          # raw CSV lines
    all_lines = []     # all output for debugging
    capturing = False
    done = False

    start = time.time()
    while not done and (time.time() - start < TIMEOUT):
        raw_bytes = robust_readline(ser, timeout_s=2.0)
        if not raw_bytes:
            continue

        raw = raw_bytes.decode("utf-8", errors="ignore").strip()
        if not raw:
            continue

        all_lines.append(raw)
        pprint(f"  {raw}")

        if raw.startswith("Freq(Hz)"):
            capturing = True
            continue

        if capturing:
            if raw.startswith("---") or raw.startswith("Sweep") or raw.startswith("Saved") or raw == "":
                if rows:
                    done = True
                    break
                continue

            parts = raw.split(",")
            if len(parts) >= 3:
                try:
                    float(parts[0])
                    rows.append(raw)
                except ValueError:
                    if rows:
                        done = True
                        break

    elapsed = time.time() - start
    ser.close()
    pprint()
    pprint(f"Capture complete. {len(rows)} data points in {elapsed:.1f}s")

    if not rows:
        pprint("No CSV data captured. Check the output above for errors.")
        return 1

    # Save to Excel
    try:
        from openpyxl import Workbook
    except ImportError:
        pprint("openpyxl not installed. Saving as CSV instead.")
        csv_file = OUT_FILE.replace(".xlsx", ".csv")
        with open(csv_file, "w") as f:
            f.write("Freq(Hz),Real(Ohm),Imag(Ohm),Magnitude(Ohm),Phase(deg)\n")
            for row in rows:
                f.write(row + "\n")
        pprint(f"Saved to {csv_file}")
        return 0

    wb = Workbook()
    ws = wb.active
    ws.title = "EIS Data"

    headers = ["Freq(Hz)", "Real(Ohm)", "Imag(Ohm)", "Magnitude(Ohm)", "Phase(deg)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, row in enumerate(rows, 2):
        parts = row.split(",")
        for col, val in enumerate(parts[:5], 1):
            try:
                ws.cell(row=row_idx, column=col, value=float(val))
            except ValueError:
                ws.cell(row=row_idx, column=col, value=val)

    wb.save(OUT_FILE)
    pprint(f"Saved {len(rows)} data points to: {os.path.abspath(OUT_FILE)}")
    pprint("Done!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
